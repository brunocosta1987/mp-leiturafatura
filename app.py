
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Comparador de Planilhas MPRJ", layout="centered")
st.title("📊 Comparador de Planilhas - Recibos MPRJ")
st.write("Compare automaticamente a planilha extraída com a planilha de referência.")

# Upload de arquivos
uploaded_extraida = st.file_uploader("📄 Envie a planilha EXTRAÍDA", type=["xlsx"])
uploaded_referencia = st.file_uploader("📄 Envie a planilha de REFERÊNCIA", type=["xlsx"])

if uploaded_extraida and uploaded_referencia:
    df_extraida = pd.read_excel(uploaded_extraida)
    df_referencia = pd.read_excel(uploaded_referencia)

    # Padronizar colunas
    df_extraida.columns = [col.strip().lower() for col in df_extraida.columns]
    df_referencia.columns = [col.strip().lower() for col in df_referencia.columns]

    df_extraida = df_extraida.rename(columns={
        'número do voucher': 'voucher',
        'valor do recibo (r$)': 'valor',
        'distância (km)': 'distancia'
    })
    df_referencia = df_referencia.rename(columns={
        'número do voucher': 'voucher',
        'valor do recibo (r$)': 'valor',
        'distância (km)': 'distancia'
    })

    # Arredondar e comparar
    df_extraida['valor'] = pd.to_numeric(df_extraida['valor'], errors='coerce').round(2)
    df_referencia['valor'] = pd.to_numeric(df_referencia['valor'], errors='coerce').round(2)
    df_extraida['distancia'] = pd.to_numeric(df_extraida['distancia'], errors='coerce').round(2)
    df_referencia['distancia'] = pd.to_numeric(df_referencia['distancia'], errors='coerce').round(2)

    df_merged = pd.merge(
        df_extraida,
        df_referencia,
        on='voucher',
        suffixes=('_extraida', '_referencia'),
        how='outer',
        indicator=True
    )

    def verificar_linha(row):
        if row['_merge'] != 'both':
            return "Voucher não encontrado"
        motivos = []
        if row['valor_extraida'] != row['valor_referencia']:
            motivos.append("Valor divergente")
        if row['distancia_extraida'] != row['distancia_referencia']:
            motivos.append("Distância divergente")
        return "Correto" if not motivos else " | ".join(motivos)

    df_merged['Status da Verificação'] = df_merged.apply(verificar_linha, axis=1)

    # Formatar visualmente os valores
    def formatar(x): return f"{x:.2f}".replace('.', ',') if pd.notnull(x) else ""
    for col in ['valor_extraida', 'valor_referencia', 'distancia_extraida', 'distancia_referencia']:
        df_merged[col] = df_merged[col].apply(formatar)

    # Salvar Excel com cor vermelha nas linhas incorretas
    excel_buffer = BytesIO()
    df_merged.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)

    wb = load_workbook(excel_buffer)
    ws = wb.active
    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    status_col = [cell.value for cell in ws[1]].index("Status da Verificação") + 1

    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=status_col).value
        if status != "Correto":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = vermelho

    final_buffer = BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)

    st.success("✅ Comparação concluída! ✅")
    st.download_button("⬇️ Baixar resultado em Excel", data=final_buffer, file_name="resultado_comparacao.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
